from sqlalchemy.orm import Session
from sqlalchemy import and_, or_
from database.models import Startup, StartupMetrics, Analysis
from typing import List, Optional
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from io import BytesIO
from datetime import datetime
import json

class ReportService:
    def __init__(self, db: Session):
        self.db = db

    def generate_startup_report(self, sectors: Optional[List[str]] = None,
                              technologies: Optional[List[str]] = None,
                              countries: Optional[List[str]] = None,
                              max_startups: int = 50,
                              sort_by: str = "score",
                              sort_order: str = "desc",
                              start_date: Optional[datetime] = None,
                              end_date: Optional[datetime] = None) -> bytes:
        """
        Gera um relatório XLSX das startups baseado nos filtros
        """
        # Buscar startups com base nos filtros
        startups = self._get_filtered_startups(sectors, technologies, countries, max_startups, sort_by, sort_order, start_date, end_date)

        # Criar workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Relatório de Startups"

        # Estilos
        header_font = Font(name='Arial', size=14, bold=True, color='1A1A1A')
        title_font = Font(name='Arial', size=16, bold=True, color='1A1A1A')
        normal_font = Font(name='Arial', size=11, color='1A1A1A')

        header_fill = PatternFill(start_color='76B900', end_color='76B900', fill_type='solid')
        info_fill = PatternFill(start_color='F0F0F0', end_color='F0F0F0', fill_type='solid')

        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        # Título principal
        ws['A1'] = "Relatório de Startups - NVIDIA Inception"
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal='center')
        ws.merge_cells('A1:H1')

        # Data de geração
        ws['A2'] = f"Gerado em: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        ws['A2'].font = normal_font
        ws.merge_cells('A2:H2')

        # Informações dos filtros
        row = 4
        ws[f'A{row}'] = "Filtros Aplicados:"
        ws[f'A{row}'].font = header_font
        row += 1

        if sectors:
            ws[f'A{row}'] = f"Setores: {', '.join(sectors)}"
            ws[f'A{row}'].font = normal_font
            row += 1

        if technologies:
            ws[f'A{row}'] = f"Tecnologias: {', '.join(technologies)}"
            ws[f'A{row}'].font = normal_font
            row += 1

        if countries:
            ws[f'A{row}'] = f"Países: {', '.join(countries)}"
            ws[f'A{row}'].font = normal_font
            row += 1

        ws[f'A{row}'] = f"Máximo de startups: {max_startups}"
        ws[f'A{row}'].font = normal_font
        row += 1

        ws[f'A{row}'] = f"Ordenado por: {sort_by} ({sort_order})"
        ws[f'A{row}'].font = normal_font
        row += 1

        if start_date or end_date:
            date_filter = ""
            if start_date:
                date_filter += f"De: {start_date.strftime('%d/%m/%Y')}"
            if end_date:
                if date_filter:
                    date_filter += f" até: {end_date.strftime('%d/%m/%Y')}"
                else:
                    date_filter += f"Até: {end_date.strftime('%d/%m/%Y')}"
            ws[f'A{row}'] = f"Período: {date_filter}"
            ws[f'A{row}'].font = normal_font
            row += 1

        ws[f'A{row}'] = f"Total encontrado: {len(startups)}"
        ws[f'A{row}'].font = normal_font
        row += 2

        # Cabeçalho da tabela
        headers = [
            'Nome', 'Setor', 'País', 'Cidade', 'Fundação', 'Website',
            'Tecnologias IA', 'Score Total', 'Score Mercado', 'Score Técnico',
            'Score Parceria', 'Funding Total (USD)', 'Último Round (USD)', 'Descrição'
        ]

        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center')

        row += 1

        # Dados das startups
        for startup in startups:
            # Buscar métricas
            metrics = startup.metrics[0] if startup.metrics else None

            data = [
                startup.name or '',
                startup.sector or '',
                startup.country or '',
                startup.city or '',
                startup.founded_year or '',
                startup.website or '',
                ', '.join(startup.ai_technologies) if startup.ai_technologies else '',
                f"{metrics.total_score:.1f}" if metrics and metrics.total_score else '0.0',
                f"{metrics.market_demand_score:.1f}" if metrics and metrics.market_demand_score else '0.0',
                f"{metrics.technical_level_score:.1f}" if metrics and metrics.technical_level_score else '0.0',
                f"{metrics.partnership_potential_score:.1f}" if metrics and metrics.partnership_potential_score else '0.0',
                f"${startup.total_funding/1000000:.1f}M" if startup.total_funding else '',
                f"${startup.last_funding_amount/1000000:.1f}M" if startup.last_funding_amount else '',
                (startup.description[:100] + '...') if startup.description and len(startup.description) > 100 else (startup.description or '')
            ]

            for col, value in enumerate(data, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

            row += 1

        # Ajustar largura das colunas
        column_widths = [25, 15, 12, 12, 10, 30, 40, 12, 12, 12, 12, 15, 15, 50]
        column_letters = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N']
        for i, width in enumerate(column_widths):
            if i < len(column_letters):
                ws.column_dimensions[column_letters[i]].width = width

        # Salvar em buffer
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        return buffer.getvalue()

    def _get_filtered_startups(self, sectors: Optional[List[str]],
                             technologies: Optional[List[str]],
                             countries: Optional[List[str]],
                             max_startups: int,
                             sort_by: str = "score",
                             sort_order: str = "desc",
                             start_date: Optional[datetime] = None,
                             end_date: Optional[datetime] = None) -> List[Startup]:
        """
        Busca startups com base nos filtros fornecidos
        """
        query = self.db.query(Startup)

        # Filtrar por setores
        if sectors:
            query = query.filter(Startup.sector.in_(sectors))

        # Filtrar por tecnologias
        if technologies:
            tech_filters = []
            for tech in technologies:
                # Use PostgreSQL JSON array contains operator
                tech_filters.append(Startup.ai_technologies.op('@>')([tech]))
            query = query.filter(or_(*tech_filters))

        # Filtrar por países
        if countries:
            query = query.filter(Startup.country.in_(countries))

        # Filtrar por data de criação
        if start_date:
            query = query.filter(Startup.created_at >= start_date)
        if end_date:
            query = query.filter(Startup.created_at <= end_date)

        # Aplicar ordenação
        query = query.outerjoin(StartupMetrics)

        if sort_by == "score":
            if sort_order == "asc":
                query = query.order_by(StartupMetrics.total_score.asc().nullslast())
            else:
                query = query.order_by(StartupMetrics.total_score.desc().nullslast())
        elif sort_by == "created_at":
            if sort_order == "asc":
                query = query.order_by(Startup.created_at.asc())
            else:
                query = query.order_by(Startup.created_at.desc())
        elif sort_by == "name":
            if sort_order == "asc":
                query = query.order_by(Startup.name.asc())
            else:
                query = query.order_by(Startup.name.desc())
        elif sort_by == "funding":
            if sort_order == "asc":
                query = query.order_by(Startup.total_funding.asc().nullslast())
            else:
                query = query.order_by(Startup.total_funding.desc().nullslast())
        else:
            # Default fallback
            query = query.order_by(
                StartupMetrics.total_score.desc().nullslast(),
                Startup.created_at.desc()
            )

        return query.limit(max_startups).all()